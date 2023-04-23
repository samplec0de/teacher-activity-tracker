from typing import List, Iterable

import openpyxl as openpyxl
from openpyxl.styles import PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from course.course import Course
from course.course_factory import CourseFactory
from lesson.lesson import Lesson
from links.course_teacher_link import CourseTeacherLink
from links.teacher_activity_link import TeacherActivityLink
from links.teacher_telegram_link import TeacherTelegramLink
from teacher.teacher import Teacher


class ReportGenerator:
    """Генератор отчета по всем курсам"""

    ALL_BORDERS = {
        'top': Side(border_style='thin'),
        'right': Side(border_style='thin'),
        'bottom': Side(border_style='thin'),
        'left': Side(border_style='thin')
    }

    FILL_COLORS = [
        PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid"),
        PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    ]

    def __init__(
            self,
            course_factory: CourseFactory,
            teacher_tg_link: TeacherTelegramLink,
            teacher_activity_link: TeacherActivityLink,
            course_teacher_link: CourseTeacherLink
    ):
        self._course_factory = course_factory
        self._teacher_tg_link = teacher_tg_link
        self._teacher_activity_link = teacher_activity_link
        self._course_teacher_link = course_teacher_link

    async def generate_report(self) -> Workbook:
        """Генерация отчета по курсам"""
        workbook = Workbook()
        courses = await self._course_factory.get_all()
        for course in courses:
            await self.add_course_sheet(workbook=workbook, course=course)
        return workbook

    async def add_course_sheet(self, workbook: Workbook, course: Course) -> None:
        """Добавление листа с отчетом по курсу

        :param workbook: Книга Excel
        :param course: Курс
        """
        sheet = workbook.active if workbook.active.title == "Sheet" else workbook.create_sheet()
        title = await course.name
        sheet.title = title

        teachers = await self._course_teacher_link.get_teachers_for_course(course)
        lessons = await course.lessons

        await self.add_teacher_info(sheet=sheet, teachers=teachers)
        await self.add_lesson_number(sheet=sheet, lessons=lessons)
        await self.add_lesson_info(sheet=sheet, lessons=lessons)
        await self.add_activity_info(sheet=sheet, lessons=lessons)
        await self.add_teacher_activity_info(sheet=sheet, teachers=teachers, lessons=lessons)

    def format_column(
            self, sheet: Worksheet, row: int, column: int, value: str, width: int
    ) -> None:
        """Форматирование ячейки в заголовке

        :param sheet: Лист Excel
        :param row: Номер строки
        :param column: Номер столбца
        :param value: Значение
        :param width: Ширина столбца
        """
        cell = sheet.cell(row=row, column=column)
        cell.value = value
        cell.font = openpyxl.styles.Font(bold=True)
        cell.alignment = openpyxl.styles.Alignment('center', wrap_text=True)
        cell.border = openpyxl.styles.Border(**self.ALL_BORDERS)
        sheet.column_dimensions[get_column_letter(column)].width = width

    async def add_teacher_info(self, sheet: Worksheet, teachers: List[Teacher]) -> None:
        """Заполнение первой строки именами преподавателей и их телеграм-логинами

        :param sheet: Лист Excel
        :param teachers: Список преподавателей
        """
        for idx, teacher in enumerate(teachers):
            column_number = idx + 4
            cell = sheet.cell(row=1, column=column_number)
            full_name = await self._teacher_tg_link.get_full_name(teacher)
            username = await self._teacher_tg_link.get_username(teacher)
            username_pretty = f"@{username}" if username else ""
            cell.value = f"{full_name}\n{username_pretty}"
            cell.alignment = openpyxl.styles.Alignment('center', wrap_text=True)
            sheet.column_dimensions[get_column_letter(column_number)].width = 18
            cell.font = openpyxl.styles.Font(bold=True)
            cell.border = openpyxl.styles.Border(**self.ALL_BORDERS)

    async def add_lesson_number(self, sheet: Worksheet, lessons: Iterable[Lesson]) -> None:
        """Заполнение первого столбца номерами уроков

        :param sheet: Лист Excel
        :param lessons: Список уроков
        """
        self.format_column(sheet=sheet, row=1, column=1, value='№', width=5)
        l_idx = 2
        for lesson_num, lesson in enumerate(lessons):
            lesson_num_cell = sheet.cell(row=l_idx, column=1)
            lesson_num_cell.value = lesson_num + 1
            lesson_num_cell.font = openpyxl.styles.Font(bold=True)
            lesson_num_cell.alignment = openpyxl.styles.Alignment(vertical='center', horizontal='center')
            lesson_num_cell.border = openpyxl.styles.Border(**self.ALL_BORDERS)
            l_idx_new = l_idx + len(await lesson.activities)
            sheet.merge_cells(start_row=l_idx, start_column=1, end_row=l_idx_new - 1, end_column=1)
            l_idx = l_idx_new

    async def add_lesson_info(self, sheet: Worksheet, lessons: Iterable[Lesson]) -> None:
        """Заполнение второго столбца темами уроков

        :param sheet: Лист Excel
        :param lessons: Список уроков
        """
        self.format_column(sheet=sheet, row=1, column=2, value='Урок', width=25)
        l_idx = 2
        for lesson_num, lesson in enumerate(lessons):
            # Заполнение темы урока во втором столбце
            lesson_cell = sheet.cell(row=l_idx, column=2)
            lesson_cell.value = await lesson.topic
            lesson_cell.font = openpyxl.styles.Font(bold=True)
            lesson_cell.alignment = openpyxl.styles.Alignment(vertical='center', wrap_text=True)
            lesson_cell.fill = self.FILL_COLORS[lesson_num % len(self.FILL_COLORS)]
            lesson_cell.border = openpyxl.styles.Border(**self.ALL_BORDERS)
            l_idx_new = l_idx + len(await lesson.activities)
            sheet.merge_cells(start_row=l_idx, start_column=2, end_row=l_idx_new - 1, end_column=2)
            l_idx = l_idx_new

    async def add_activity_info(self, sheet: Worksheet, lessons: Iterable[Lesson]) -> None:
        """Заполнение третьего столбца названиями активностей

        :param sheet: Лист Excel
        :param lessons: Список уроков
        """
        self.format_column(sheet=sheet, row=1, column=3, value='Активность', width=25)
        l_idx = 2
        for lesson_num, lesson in enumerate(lessons):
            activities = await lesson.activities
            for a_idx, activity in enumerate(activities):
                activity_cell = sheet.cell(row=l_idx + a_idx, column=3)
                activity_cell.value = await activity.name
                activity_cell.alignment = openpyxl.styles.Alignment(wrap_text=True)
                activity_cell.border = openpyxl.styles.Border(**self.ALL_BORDERS)
            l_idx += len(await lesson.activities)

    async def add_teacher_activity_info(
            self, sheet: Worksheet, lessons: Iterable[Lesson], teachers: List[Teacher]
    ) -> None:
        """Заполнение пересечения учителя и активности количеством часов

        :param sheet: Лист Excel
        :param lessons: Список уроков
        :param teachers: Список преподавателей
        """
        l_idx = 2
        for lesson_num, lesson in enumerate(lessons):

            activities = await lesson.activities
            for a_idx, activity in enumerate(activities):
                for idx, teacher in enumerate(teachers):
                    hours = await self._teacher_activity_link.get_hours_for_activity_and_teacher(activity, teacher)
                    activity_hours_cell = sheet.cell(row=l_idx + a_idx, column=idx + 4)
                    activity_hours_cell.border = openpyxl.styles.Border(**self.ALL_BORDERS)
                    activity_hours_cell.alignment = openpyxl.styles.Alignment(vertical='center', horizontal='center')
                    if hours:
                        activity_hours_cell.value = hours

            l_idx += len(activities)


if __name__ == '__main__':
    import asyncio

    from links.teacher_activity_link import TeacherActivityLink, PGTeacherActivityLink
    from database import get_pool

    import bot

    async def main():
        generator = ReportGenerator(
            course_factory=await bot.get_course_factory(),
            teacher_tg_link=TeacherTelegramLink(bot=bot.bot),
            teacher_activity_link=PGTeacherActivityLink(await get_pool()),
            course_teacher_link=CourseTeacherLink(await get_pool()),
        )

        workbook: Workbook = await generator.generate_report()
        workbook.save('report.xlsx')

    asyncio.run(main())
