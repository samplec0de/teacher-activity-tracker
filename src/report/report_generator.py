from typing import List

import openpyxl as openpyxl
from openpyxl.styles import PatternFill, Side

from course.course import Course
from links.teacher_activity_link import TeacherActivityLink
from links.teacher_telegram_link import TeacherTelegramLink
from teacher.teacher import Teacher


class ReportGenerator:
    def __init__(self, teacher_tg_link: TeacherTelegramLink, teacher_activity_link: TeacherActivityLink):
        self._teacher_tg_link = teacher_tg_link
        self._teacher_activity_link = teacher_activity_link

    async def generate_report(self, courses: List[Course], teachers: List[Teacher], file_name: str):
        # Создание новой книги
        fill_colors = [
            PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid"),
            PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        ]
        thin_border = Side(border_style='thin')
        all_border = {
            'top': thin_border,
            'right': thin_border,
            'bottom': thin_border,
            'left': thin_border
        }
        workbook = openpyxl.Workbook()
        for course in courses:
            # Создание нового листа с именем курса
            sheet = workbook.active if workbook.active.title == "Sheet" else workbook.create_sheet()
            title = await course.name
            sheet.title = title

            lesson_num_title = sheet.cell(row=1, column=1)
            lesson_num_title.value = '№'
            lesson_num_title.font = openpyxl.styles.Font(bold=True)
            lesson_num_title.alignment = openpyxl.styles.Alignment('center', wrap_text=True)
            lesson_num_title.border = openpyxl.styles.Border(**all_border)
            sheet.column_dimensions['A'].width = 5

            lesson_title = sheet.cell(row=1, column=2)
            lesson_title.value = 'Урок'
            lesson_title.font = openpyxl.styles.Font(bold=True)
            lesson_title.alignment = openpyxl.styles.Alignment('center', wrap_text=True)
            lesson_title.border = openpyxl.styles.Border(**all_border)
            sheet.column_dimensions['B'].width = 25

            activity_title = sheet.cell(row=1, column=3)
            activity_title.value = 'Активность'
            activity_title.font = openpyxl.styles.Font(bold=True)
            activity_title.alignment = openpyxl.styles.Alignment('center', wrap_text=True)
            activity_title.border = openpyxl.styles.Border(**all_border)
            sheet.column_dimensions['C'].width = 25

            # Заполнение первой строки именами преподавателей
            for idx, teacher in enumerate(teachers):
                column_number = idx + 4
                cell = sheet.cell(row=1, column=column_number)
                cell.value = await self._teacher_tg_link.get_full_name(teacher)
                cell.alignment = openpyxl.styles.Alignment('center', wrap_text=True)
                sheet.column_dimensions[self.get_excel_column_name(column_number)].width = 18
                cell.font = openpyxl.styles.Font(bold=True)
                cell.border = openpyxl.styles.Border(**all_border)

            l_idx = 2
            lessons = await course.lessons
            for lesson_num, lesson in enumerate(lessons):
                # Заполнение списка уроков курса в первом столбце
                lesson_num_cell = sheet.cell(row=l_idx, column=1)
                lesson_num_cell.value = lesson_num + 1
                lesson_num_cell.font = openpyxl.styles.Font(bold=True)
                lesson_num_cell.alignment = openpyxl.styles.Alignment(vertical='center', horizontal='center')
                lesson_num_cell.border = openpyxl.styles.Border(**all_border)

                lesson_cell = sheet.cell(row=l_idx, column=2)
                lesson_cell.value = await lesson.topic
                lesson_cell.font = openpyxl.styles.Font(bold=True)
                lesson_cell.alignment = openpyxl.styles.Alignment(vertical='center', wrap_text=True)
                lesson_cell.fill = fill_colors[lesson_num % len(fill_colors)]
                lesson_cell.border = openpyxl.styles.Border(**all_border)

                activities = await lesson.activities
                for a_idx, activity in enumerate(activities):
                    # Заполнение списка активностей во втором столбце
                    activity_cell = sheet.cell(row=l_idx + a_idx, column=3)
                    activity_cell.value = await activity.name
                    activity_cell.alignment = openpyxl.styles.Alignment(wrap_text=True)
                    activity_cell.border = openpyxl.styles.Border(**all_border)

                    # Заполнение ячеек пересечения активности и учителя количеством часов
                    for idx, teacher in enumerate(teachers):
                        hours = await self._teacher_activity_link.get_hours_for_activity_and_teacher(activity, teacher)
                        activity_hours_cell = sheet.cell(row=l_idx + a_idx, column=idx + 4)
                        activity_hours_cell.border = openpyxl.styles.Border(**all_border)
                        activity_hours_cell.alignment = openpyxl.styles.Alignment(vertical='center', horizontal='center')
                        if hours:
                            activity_hours_cell.value = hours

                l_idx_new = l_idx + len(await lesson.activities)
                # Объединение ячеек урока
                sheet.merge_cells(start_row=l_idx, start_column=2, end_row=l_idx_new - 1, end_column=2)
                sheet.merge_cells(start_row=l_idx, start_column=1, end_row=l_idx_new - 1, end_column=1)
                l_idx = l_idx_new

        # Сохранение файла
        workbook.save(file_name)

    @staticmethod
    def get_excel_column_name(column_number: int) -> str:
        """Получение имени столбца Excel по его номеру"""
        column_name = ""
        while column_number > 0:
            column_number, remainder = divmod(column_number - 1, 26)
            column_name = chr(65 + remainder) + column_name
        return column_name


if __name__ == '__main__':
    import asyncio

    from links.teacher_activity_link import TeacherActivityLink, PGTeacherActivityLink
    from database import get_pool

    import bot

    async def main():
        generator = ReportGenerator(
            teacher_tg_link=TeacherTelegramLink(bot=bot.bot),
            teacher_activity_link=PGTeacherActivityLink(await get_pool())
        )
        cf = await bot.get_course_factory()
        tf = await bot.get_teacher_factory()
        await generator.generate_report(
            courses=await cf.get_all(),
            teachers=await tf.get_all(),
            file_name='report.xlsx'
        )

    asyncio.run(main())
